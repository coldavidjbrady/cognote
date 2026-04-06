#!/usr/bin/env python3
"""
Apple Notes streaming exporter for macOS.

Why this version exists:
- avoids building one giant JSON blob in AppleScript
- processes notes folder-by-folder
- writes CSV / JSONL / Markdown incrementally
- writes XLSX in write-only mode (streaming) when openpyxl is installed
- supports resume/checkpointing for large exports

Outputs:
- notes_export.csv
- notes_export.jsonl
- notes_merged.md
- notes_export.xlsx        (optional; requires openpyxl)
- export_summary.json
- .export_state.json       (only when --resume is used)

This script uses AppleScript via osascript to talk to the Notes app.
It must be run on macOS with Automation permission granted for the terminal
or Python app that executes it.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
import subprocess
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterator, List, Optional, Tuple


US = chr(31)  # field separator
RS = chr(30)  # record separator
GS = chr(29)  # folder-path segment separator
ILLEGAL_XLSX_CHAR_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


LIST_FOLDERS_SCRIPT = r'''
use scripting additions

property fieldSep : character id 31
property recordSep : character id 30
property pathSep : character id 29

on esc(theText)
    if theText is missing value then set theText to ""
    set t to theText as text

    set AppleScript's text item delimiters to "\\"
    set t to every text item of t
    set AppleScript's text item delimiters to "\\\\"
    set t to t as text

    set AppleScript's text item delimiters to fieldSep
    set t to every text item of t
    set AppleScript's text item delimiters to "\\u001f"
    set t to t as text

    set AppleScript's text item delimiters to recordSep
    set t to every text item of t
    set AppleScript's text item delimiters to "\\u001e"
    set t to t as text

    set AppleScript's text item delimiters to pathSep
    set t to every text item of t
    set AppleScript's text item delimiters to "\\u001d"
    set t to t as text

    set AppleScript's text item delimiters to "\""
    set t to every text item of t
    set AppleScript's text item delimiters to "\\\""
    set t to t as text

    set AppleScript's text item delimiters to return
    set t to every text item of t
    set AppleScript's text item delimiters to "\\n"
    set t to t as text

    set AppleScript's text item delimiters to linefeed
    set t to every text item of t
    set AppleScript's text item delimiters to "\\n"
    set t to t as text

    set AppleScript's text item delimiters to tab
    set t to every text item of t
    set AppleScript's text item delimiters to "\\t"
    set t to t as text

    set AppleScript's text item delimiters to ""
    return t
end esc

on joinPath(pathText, folderName)
    if pathText is "" then
        return folderName
    end if
    return pathText & pathSep & folderName
end joinPath

on walkFolders(folderList, accName, currentPath, output)
    repeat with f in folderList
        set folderName to ""
        try
            set folderName to name of f as text
        end try

        set nextPath to my joinPath(currentPath, folderName)
        set output to output & my esc(accName) & fieldSep & my esc(nextPath) & recordSep

        try
            set output to my walkFolders(folders of f, accName, nextPath, output)
        end try
    end repeat
    return output
end walkFolders

tell application "Notes"
    set output to ""
    repeat with acc in accounts
        set accName to ""
        try
            set accName to name of acc as text
        end try
        set output to my walkFolders(folders of acc, accName, "", output)
    end repeat
end tell

return output
'''


FETCH_FOLDER_TEMPLATE = r'''
use scripting additions

property fieldSep : character id 31
property recordSep : character id 30
property pathSep : character id 29

on esc(theText)
    if theText is missing value then set theText to ""
    set t to theText as text

    set AppleScript's text item delimiters to "\\"
    set t to every text item of t
    set AppleScript's text item delimiters to "\\\\"
    set t to t as text

    set AppleScript's text item delimiters to fieldSep
    set t to every text item of t
    set AppleScript's text item delimiters to "\\u001f"
    set t to t as text

    set AppleScript's text item delimiters to recordSep
    set t to every text item of t
    set AppleScript's text item delimiters to "\\u001e"
    set t to t as text

    set AppleScript's text item delimiters to pathSep
    set t to every text item of t
    set AppleScript's text item delimiters to "\\u001d"
    set t to t as text

    set AppleScript's text item delimiters to "\""
    set t to every text item of t
    set AppleScript's text item delimiters to "\\\""
    set t to t as text

    set AppleScript's text item delimiters to return
    set t to every text item of t
    set AppleScript's text item delimiters to "\\n"
    set t to t as text

    set AppleScript's text item delimiters to linefeed
    set t to every text item of t
    set AppleScript's text item delimiters to "\\n"
    set t to t as text

    set AppleScript's text item delimiters to tab
    set t to every text item of t
    set AppleScript's text item delimiters to "\\t"
    set t to t as text

    set AppleScript's text item delimiters to ""
    return t
end esc

on folderFromPath(accountName, encodedPath)
    set AppleScript's text item delimiters to pathSep
    set parts to every text item of encodedPath
    set AppleScript's text item delimiters to ""

    tell application "Notes"
        set accRef to first account whose name is accountName
        set currentFolder to missing value
        set currentList to folders of accRef

        repeat with partName in parts
            set targetName to partName as text
            set foundFolder to missing value

            repeat with f in currentList
                try
                    if (name of f as text) is targetName then
                        set foundFolder to f
                        exit repeat
                    end if
                end try
            end repeat

            if foundFolder is missing value then error "Folder path not found: " & encodedPath

            set currentFolder to foundFolder
            try
                set currentList to folders of currentFolder
            on error
                set currentList to {}
            end try
        end repeat

        return currentFolder
    end tell
end folderFromPath

set accountName to __ACCOUNT_NAME__
set folderPath to __FOLDER_PATH__

tell application "Notes"
    set output to ""
    set folderRef to my folderFromPath(accountName, folderPath)

    repeat with n in notes of folderRef
        try
            set noteName to name of n as text
        on error
            set noteName to ""
        end try

        try
            set noteBody to body of n as text
        on error
            set noteBody to ""
        end try

        try
            set creationDate to creation date of n as text
        on error
            set creationDate to ""
        end try

        try
            set modificationDate to modification date of n as text
        on error
            set modificationDate to ""
        end try

        try
            set noteID to id of n as text
        on error
            set noteID to ""
        end try

        set output to output & my esc(accountName) & fieldSep & my esc(folderPath) & fieldSep & my esc(noteID) & fieldSep & my esc(noteName) & fieldSep & my esc(creationDate) & fieldSep & my esc(modificationDate) & fieldSep & my esc(noteBody) & recordSep
    end repeat
end tell

return output
'''


FETCH_ALL_NOTES_TEMPLATE = r'''
use scripting additions

property fieldSep : character id 31
property recordSep : character id 30
property pathSep : character id 29

on esc(theText)
    if theText is missing value then set theText to ""
    set t to theText as text

    set AppleScript's text item delimiters to "\\"
    set t to every text item of t
    set AppleScript's text item delimiters to "\\\\"
    set t to t as text

    set AppleScript's text item delimiters to fieldSep
    set t to every text item of t
    set AppleScript's text item delimiters to "\\u001f"
    set t to t as text

    set AppleScript's text item delimiters to recordSep
    set t to every text item of t
    set AppleScript's text item delimiters to "\\u001e"
    set t to t as text

    set AppleScript's text item delimiters to pathSep
    set t to every text item of t
    set AppleScript's text item delimiters to "\\u001d"
    set t to t as text

    set AppleScript's text item delimiters to "\""
    set t to every text item of t
    set AppleScript's text item delimiters to "\\\""
    set t to t as text

    set AppleScript's text item delimiters to return
    set t to every text item of t
    set AppleScript's text item delimiters to "\\n"
    set t to t as text

    set AppleScript's text item delimiters to linefeed
    set t to every text item of t
    set AppleScript's text item delimiters to "\\n"
    set t to t as text

    set AppleScript's text item delimiters to tab
    set t to every text item of t
    set AppleScript's text item delimiters to "\\t"
    set t to t as text

    set AppleScript's text item delimiters to ""
    return t
end esc

on joinPath(pathText, folderName)
    if pathText is "" then
        return folderName
    end if
    return pathText & pathSep & folderName
end joinPath

on appendNotes(folderRef, accName, folderPath, output)
    repeat with n in notes of folderRef
        try
            set noteName to name of n as text
        on error
            set noteName to ""
        end try

        try
            set noteBody to body of n as text
        on error
            set noteBody to ""
        end try

        try
            set creationDate to creation date of n as text
        on error
            set creationDate to ""
        end try

        try
            set modificationDate to modification date of n as text
        on error
            set modificationDate to ""
        end try

        try
            set noteID to id of n as text
        on error
            set noteID to ""
        end try

        set output to output & my esc(accName) & fieldSep & my esc(folderPath) & fieldSep & my esc(noteID) & fieldSep & my esc(noteName) & fieldSep & my esc(creationDate) & fieldSep & my esc(modificationDate) & fieldSep & my esc(noteBody) & recordSep
    end repeat
    return output
end appendNotes

on walkFolders(folderList, accName, currentPath, output)
    repeat with f in folderList
        set folderName to ""
        try
            set folderName to name of f as text
        end try

        set nextPath to my joinPath(currentPath, folderName)
        set output to my appendNotes(f, accName, nextPath, output)
        try
            set output to my walkFolders(folders of f, accName, nextPath, output)
        end try
    end repeat
    return output
end walkFolders

set accountNameFilter to __ACCOUNT_NAME_FILTER__

tell application "Notes"
    set output to ""
    repeat with acc in accounts
        set accName to ""
        try
            set accName to name of acc as text
        end try

        if accountNameFilter is "" or accName is accountNameFilter then
            set output to my walkFolders(folders of acc, accName, "", output)
        end if
    end repeat
end tell

return output
'''


@dataclass
class NoteRecord:
    account: str
    folder: str
    id: str
    title: str
    created: str
    modified: str
    body_html: str
    body_text: str
    word_count: int
    char_count: int


def applescript_string_literal(value: str) -> str:
    return '"' + value.replace("\\", "\\\\").replace('"', '\\"') + '"'


def run_osascript(script: str) -> str:
    proc = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        raise RuntimeError(
            "osascript failed. Make sure you are on macOS and have granted "
            "Automation access to control Notes.\n\n"
            f"stderr:\n{proc.stderr.strip()}"
        )
    return proc.stdout


def unescape_applescript_field(value: str) -> str:
    return (
        value.replace("\\u001f", US)
        .replace("\\u001e", RS)
        .replace("\\u001d", GS)
        .replace("\\t", "\t")
        .replace("\\n", "\n")
        .replace("\\\"", '"')
        .replace("\\\\", "\\")
    )


def strip_html_simple(html_text: str) -> str:
    if not html_text:
        return ""
    text = re.sub(r"(?i)<br\s*/?>", "\n", html_text)
    text = re.sub(r"(?i)</p\s*>", "\n\n", text)
    text = re.sub(r"(?i)</div\s*>", "\n", text)
    text = re.sub(r"(?i)</li\s*>", "\n", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = html.unescape(text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def parse_folder_listing(raw: str) -> List[Tuple[str, str]]:
    items: List[Tuple[str, str]] = []
    for rec in raw.split(RS):
        if not rec:
            continue
        parts = rec.split(US)
        if len(parts) != 2:
            continue
        account = unescape_applescript_field(parts[0])
        folder_path = unescape_applescript_field(parts[1])
        if account and folder_path:
            items.append((account, folder_path))
    return items


def parse_note_records(raw: str) -> Iterator[NoteRecord]:
    for rec in raw.split(RS):
        if not rec:
            continue
        parts = rec.split(US)
        if len(parts) != 7:
            continue
        account, folder_encoded, note_id, title, created, modified, body_html = [
            unescape_applescript_field(p) for p in parts
        ]
        folder_display = folder_encoded.replace(GS, " / ")
        body_text = strip_html_simple(body_html)
        word_count = len([w for w in body_text.split() if w.strip()])
        yield NoteRecord(
            account=account,
            folder=folder_display,
            id=note_id,
            title=title,
            created=created,
            modified=modified,
            body_html=body_html,
            body_text=body_text,
            word_count=word_count,
            char_count=len(body_text),
        )


def fetch_folders(account_filter: Optional[str]) -> List[Tuple[str, str]]:
    folders = parse_folder_listing(run_osascript(LIST_FOLDERS_SCRIPT))
    if account_filter:
        folders = [item for item in folders if item[0] == account_filter]
    return folders


def fetch_folder_notes(account: str, folder_path: str) -> Iterator[NoteRecord]:
    script = (
        FETCH_FOLDER_TEMPLATE
        .replace("__ACCOUNT_NAME__", applescript_string_literal(account))
        .replace("__FOLDER_PATH__", applescript_string_literal(folder_path))
    )
    raw = run_osascript(script)
    return parse_note_records(raw)


def fetch_all_notes(account_filter: Optional[str]) -> List[NoteRecord]:
    account_name_filter = account_filter or ""
    script = FETCH_ALL_NOTES_TEMPLATE.replace(
        "__ACCOUNT_NAME_FILTER__",
        applescript_string_literal(account_name_filter),
    )
    raw = run_osascript(script)
    return list(parse_note_records(raw))


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def load_state(path: Path) -> Dict[str, object]:
    if not path.exists():
        return {"completed_folders": [], "stats": {}}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {"completed_folders": [], "stats": {}}


def save_state(path: Path, completed_folders: List[str], stats: Dict[str, object]) -> None:
    payload = {
        "updated_at": datetime.now().isoformat(),
        "completed_folders": completed_folders,
        "stats": stats,
    }
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def write_csv_records(notes: List[NoteRecord], path: Path) -> None:
    with path.open("w", newline="", encoding="utf-8") as csv_file:
        csv_writer = csv.DictWriter(
            csv_file,
            fieldnames=[
                "account",
                "folder",
                "id",
                "title",
                "created",
                "modified",
                "body_html",
                "body_text",
                "word_count",
                "char_count",
            ],
        )
        csv_writer.writeheader()
        for note in notes:
            csv_writer.writerow(asdict(note))


def write_jsonl_records(notes: List[NoteRecord], path: Path) -> None:
    with path.open("w", encoding="utf-8") as jsonl_file:
        for note in notes:
            jsonl_file.write(json.dumps(asdict(note), ensure_ascii=False) + "\n")


def write_markdown_records(notes: List[NoteRecord], path: Path) -> None:
    with path.open("w", encoding="utf-8") as md_file:
        md_file.write("# Apple Notes Export\n\n")
        md_file.write(f"Export started: {datetime.now().isoformat()}\n\n")
        for note in notes:
            md_file.write(f"---\n\n## {note.title or '(Untitled)'}\n\n")
            md_file.write(f"- Account: {note.account}\n")
            md_file.write(f"- Folder: {note.folder}\n")
            md_file.write(f"- ID: {note.id}\n")
            md_file.write(f"- Created: {note.created}\n")
            md_file.write(f"- Modified: {note.modified}\n")
            md_file.write(f"- Words: {note.word_count}\n\n")
            if note.body_text:
                md_file.write(note.body_text)
                md_file.write("\n\n")


def write_xlsx_records(notes: List[NoteRecord], path: Path) -> bool:
    writer = StreamingXlsxWriter(path)
    if not writer.enabled:
        return False
    for note in notes:
        writer.append(note)
    writer.finalize(build_stats(notes, path.parent))
    return True


def build_stats(notes: List[NoteRecord], out_dir: Path) -> Dict[str, object]:
    account_names = sorted({note.account for note in notes})
    folder_names = sorted({note.folder for note in notes})
    return {
        "exported_at": datetime.now().isoformat(),
        "total_notes": len(notes),
        "total_words": sum(note.word_count for note in notes),
        "total_chars": sum(note.char_count for note in notes),
        "total_accounts": len(account_names),
        "total_folders": len(folder_names),
        "account_names": account_names,
        "folder_names": folder_names,
        "output_dir": str(out_dir),
    }


def sanitize_for_xlsx(value):
    if isinstance(value, str):
        return ILLEGAL_XLSX_CHAR_RE.sub("", value)
    return value


class StreamingXlsxWriter:
    def __init__(self, path: Path):
        self.path = path
        self.enabled = False
        self._wb = None
        self._ws = None
        self._summary = None
        try:
            from openpyxl import Workbook  # type: ignore
            self._wb = Workbook(write_only=True)
            self._ws = self._wb.create_sheet("notes")
            self._ws.append([
                "account",
                "folder",
                "id",
                "title",
                "created",
                "modified",
                "word_count",
                "char_count",
                "body_text",
                "body_html",
            ])
            self._summary = self._wb.create_sheet("summary")
            self.enabled = True
        except Exception:
            self.enabled = False

    def append(self, note: NoteRecord) -> None:
        if not self.enabled:
            return
        self._ws.append([
            sanitize_for_xlsx(note.account),
            sanitize_for_xlsx(note.folder),
            sanitize_for_xlsx(note.id),
            sanitize_for_xlsx(note.title),
            sanitize_for_xlsx(note.created),
            sanitize_for_xlsx(note.modified),
            note.word_count,
            note.char_count,
            sanitize_for_xlsx(note.body_text),
            sanitize_for_xlsx(note.body_html),
        ])

    def finalize(self, stats: Dict[str, object]) -> None:
        if not self.enabled:
            return
        self._summary.append(["metric", "value"])
        self._summary.append(["exported_at", sanitize_for_xlsx(datetime.now().isoformat())])
        for key in [
            "total_notes",
            "total_accounts",
            "total_folders",
            "total_words",
            "total_chars",
        ]:
            self._summary.append([sanitize_for_xlsx(key), sanitize_for_xlsx(stats.get(key, 0))])
        self._wb.save(self.path)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Stream-export Apple Notes into merged files.")
    p.add_argument("--output-dir", required=True, help="Directory for export outputs")
    p.add_argument("--account", default=None, help="Optional exact account name filter, e.g. iCloud")
    p.add_argument("--skip-xlsx", action="store_true", help="Skip XLSX generation")
    p.add_argument("--resume", action="store_true", help="Resume from .export_state.json if it exists")
    p.add_argument("--state-file", default=".export_state.json", help="State filename inside output dir")
    p.add_argument("--progress-every", type=int, default=25, help="Print progress every N notes")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    if sys.platform != "darwin":
        print("This script only runs on macOS.", file=sys.stderr)
        return 2

    args = parse_args(argv)
    out_dir = Path(os.path.expanduser(args.output_dir)).resolve()
    ensure_dir(out_dir)

    state_path = out_dir / args.state_file
    state = load_state(state_path) if args.resume else {"completed_folders": [], "stats": {}}
    completed_folders = set(state.get("completed_folders", []))

    folders = fetch_folders(args.account)
    if not folders:
        print("No folders found for the selected scope.")
        return 0

    csv_path = out_dir / "notes_export.csv"
    jsonl_path = out_dir / "notes_export.jsonl"
    md_path = out_dir / "notes_merged.md"
    xlsx_path = out_dir / "notes_export.xlsx"
    summary_path = out_dir / "export_summary.json"

    csv_exists = csv_path.exists() and args.resume
    md_exists = md_path.exists() and args.resume

    csv_file = csv_path.open("a" if args.resume else "w", newline="", encoding="utf-8")
    jsonl_file = jsonl_path.open("a" if args.resume else "w", encoding="utf-8")
    md_file = md_path.open("a" if args.resume else "w", encoding="utf-8")

    csv_writer = csv.DictWriter(
        csv_file,
        fieldnames=[
            "account",
            "folder",
            "id",
            "title",
            "created",
            "modified",
            "body_html",
            "body_text",
            "word_count",
            "char_count",
        ],
    )
    if not csv_exists:
        csv_writer.writeheader()
    if not md_exists:
        md_file.write("# Apple Notes Export\n\n")
        md_file.write(f"Export started: {datetime.now().isoformat()}\n\n")

    xlsx_writer = None if args.skip_xlsx else StreamingXlsxWriter(xlsx_path)

    total_notes = 0
    total_words = 0
    total_chars = 0
    account_names = set()
    folder_names = set()

    if args.resume and isinstance(state.get("stats"), dict):
        prior_stats = state["stats"]
        total_notes = int(prior_stats.get("total_notes", 0))
        total_words = int(prior_stats.get("total_words", 0))
        total_chars = int(prior_stats.get("total_chars", 0))
        account_names = set(prior_stats.get("account_names", []))
        folder_names = set(prior_stats.get("folder_names", []))

    try:
        for idx, (account, folder_path) in enumerate(folders, start=1):
            folder_key = f"{account}{US}{folder_path}"
            if args.resume and folder_key in completed_folders:
                print(f"[skip {idx}/{len(folders)}] {account} :: {folder_path.replace(GS, ' / ')}")
                continue

            display_folder = folder_path.replace(GS, " / ")
            print(f"[folder {idx}/{len(folders)}] {account} :: {display_folder}")

            folder_note_count = 0
            for note in fetch_folder_notes(account, folder_path):
                row = asdict(note)
                csv_writer.writerow(row)
                jsonl_file.write(json.dumps(row, ensure_ascii=False) + "\n")

                md_file.write(f"---\n\n## {note.title or '(Untitled)'}\n\n")
                md_file.write(f"- Account: {note.account}\n")
                md_file.write(f"- Folder: {note.folder}\n")
                md_file.write(f"- ID: {note.id}\n")
                md_file.write(f"- Created: {note.created}\n")
                md_file.write(f"- Modified: {note.modified}\n")
                md_file.write(f"- Words: {note.word_count}\n\n")
                if note.body_text:
                    md_file.write(note.body_text)
                    md_file.write("\n\n")

                if xlsx_writer and xlsx_writer.enabled:
                    xlsx_writer.append(note)

                total_notes += 1
                total_words += note.word_count
                total_chars += note.char_count
                account_names.add(note.account)
                folder_names.add(note.folder)
                folder_note_count += 1

                if args.progress_every > 0 and total_notes % args.progress_every == 0:
                    print(f"  processed {total_notes} notes...")

            completed_folders.add(folder_key)
            if args.resume:
                save_state(
                    state_path,
                    sorted(completed_folders),
                    {
                        "total_notes": total_notes,
                        "total_words": total_words,
                        "total_chars": total_chars,
                        "total_accounts": len(account_names),
                        "total_folders": len(folder_names),
                        "account_names": sorted(account_names),
                        "folder_names": sorted(folder_names),
                    },
                )
            print(f"  finished folder with {folder_note_count} notes")
    finally:
        csv_file.close()
        jsonl_file.close()
        md_file.close()

    if total_notes == 0 and folders:
        print("Streaming export found folders but no notes. Retrying with bulk fallback...")
        fallback_notes = fetch_all_notes(args.account)
        if fallback_notes:
            write_csv_records(fallback_notes, csv_path)
            write_jsonl_records(fallback_notes, jsonl_path)
            write_markdown_records(fallback_notes, md_path)

            xlsx_enabled = False
            if not args.skip_xlsx:
                xlsx_enabled = write_xlsx_records(fallback_notes, xlsx_path)

            stats = build_stats(fallback_notes, out_dir)
            stats["xlsx_enabled"] = xlsx_enabled
            summary_path.write_text(json.dumps(stats, indent=2), encoding="utf-8")
            if args.resume:
                completed_folder_keys = [f"{account}{US}{folder_path}" for account, folder_path in folders]
                save_state(state_path, sorted(completed_folder_keys), stats)

            print(f"Fallback exporter recovered {len(fallback_notes)} notes")
            print("\nExport complete")
            print(f"CSV:     {csv_path}")
            print(f"JSONL:   {jsonl_path}")
            print(f"MD:      {md_path}")
            print(f"SUMMARY: {summary_path}")
            if args.resume:
                print(f"STATE:   {state_path}")
            if xlsx_enabled:
                print(f"XLSX:    {xlsx_path}")
            elif not args.skip_xlsx:
                print("XLSX skipped because openpyxl is not available.")
            return 0

    stats = {
        "exported_at": datetime.now().isoformat(),
        "total_notes": total_notes,
        "total_words": total_words,
        "total_chars": total_chars,
        "total_accounts": len(account_names),
        "total_folders": len(folder_names),
        "account_names": sorted(account_names),
        "folder_names": sorted(folder_names),
        "output_dir": str(out_dir),
        "xlsx_enabled": bool(xlsx_writer and xlsx_writer.enabled),
    }
    summary_path.write_text(json.dumps(stats, indent=2), encoding="utf-8")

    if xlsx_writer and xlsx_writer.enabled:
        print("Saving XLSX workbook...")
        xlsx_writer.finalize(stats)
    elif not args.skip_xlsx:
        print("XLSX skipped because openpyxl is not available.")

    print("\nExport complete")
    print(f"CSV:     {csv_path}")
    print(f"JSONL:   {jsonl_path}")
    print(f"MD:      {md_path}")
    print(f"SUMMARY: {summary_path}")
    if args.resume:
        print(f"STATE:   {state_path}")
    if xlsx_writer and xlsx_writer.enabled:
        print(f"XLSX:    {xlsx_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
