#!/usr/bin/env python3
"""
Apple Notes bulk exporter for macOS.

Exports note text and metadata from Apple Notes into:
- XLSX workbook (optional; requires openpyxl)
- CSV
- JSONL
- merged Markdown

This script uses AppleScript via osascript to talk to the Notes app.
It is intended to run on macOS with Automation permission granted.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import subprocess
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Optional


APPLE_SCRIPT = r'''
use scripting additions

on esc(theText)
    set t to theText as text

    set AppleScript's text item delimiters to "\\"
    set t to every text item of t
    set AppleScript's text item delimiters to "\\\\"
    set t to t as text

    set AppleScript's text item delimiters to "\""
    set t to every text item of t
    set AppleScript's text item delimiters to "\\\""
    set t to t as text

    set AppleScript's text item delimiters to return
    set t to every text item of t
    set AppleScript's text item delimiters to "\\n"
    set t to t as text

    set AppleScript's text item delimiters to linefeed
    set t to every text item of t
    set AppleScript's text item delimiters to "\\n"
    set t to t as text

    set AppleScript's text item delimiters to tab
    set t to every text item of t
    set AppleScript's text item delimiters to "\\t"
    set t to t as text

    set AppleScript's text item delimiters to ""
    return t
end esc

on nullToEmpty(v)
    if v is missing value then
        return ""
    end if
    return v as text
end nullToEmpty

tell application "Notes"
    set output to "["
    set firstRecord to true

    repeat with acc in accounts
        set accName to name of acc as text
        repeat with f in folders of acc
            set folderName to name of f as text
            repeat with n in notes of f
                try
                    set noteName to name of n as text
                on error
                    set noteName to ""
                end try

                try
                    set noteBody to body of n as text
                on error
                    set noteBody to ""
                end try

                try
                    set creationDate to creation date of n as text
                on error
                    set creationDate to ""
                end try

                try
                    set modificationDate to modification date of n as text
                on error
                    set modificationDate to ""
                end try

                try
                    set noteID to id of n as text
                on error
                    set noteID to ""
                end try

                if firstRecord is false then
                    set output to output & ","
                end if
                set firstRecord to false

                set output to output & "{"
                set output to output & "\"account\":\"" & my esc(accName) & "\"," 
                set output to output & "\"folder\":\"" & my esc(folderName) & "\"," 
                set output to output & "\"id\":\"" & my esc(noteID) & "\"," 
                set output to output & "\"title\":\"" & my esc(noteName) & "\"," 
                set output to output & "\"created\":\"" & my esc(creationDate) & "\"," 
                set output to output & "\"modified\":\"" & my esc(modificationDate) & "\"," 
                set output to output & "\"body_html\":\"" & my esc(noteBody) & "\""
                set output to output & "}"
            end repeat
        end repeat
    end repeat

    set output to output & "]"
end tell

return output
'''


@dataclass
class NoteRecord:
    account: str
    folder: str
    id: str
    title: str
    created: str
    modified: str
    body_html: str
    body_text: str
    word_count: int
    char_count: int


def strip_html_simple(html: str) -> str:
    import re

    if not html:
        return ""
    text = re.sub(r"<br\\s*/?>", "\n", html, flags=re.IGNORECASE)
    text = re.sub(r"</p\\s*>", "\n\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = (
        text.replace("&nbsp;", " ")
        .replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&quot;", '"')
        .replace("&#39;", "'")
    )
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def run_osascript(script: str) -> str:
    proc = subprocess.run(
        ["osascript", "-e", script],
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        raise RuntimeError(
            "osascript failed. Make sure you are on macOS and have granted "
            "Automation access to control Notes.\n\n"
            f"stderr:\n{proc.stderr.strip()}"
        )
    return proc.stdout.strip()


def fetch_notes(account_filter: Optional[str] = None) -> List[NoteRecord]:
    raw = run_osascript(APPLE_SCRIPT)
    items = json.loads(raw)
    notes: List[NoteRecord] = []
    for item in items:
        if account_filter and item.get("account") != account_filter:
            continue
        body_text = strip_html_simple(item.get("body_html", ""))
        words = len([w for w in body_text.split() if w.strip()])
        notes.append(
            NoteRecord(
                account=item.get("account", ""),
                folder=item.get("folder", ""),
                id=item.get("id", ""),
                title=item.get("title", ""),
                created=item.get("created", ""),
                modified=item.get("modified", ""),
                body_html=item.get("body_html", ""),
                body_text=body_text,
                word_count=words,
                char_count=len(body_text),
            )
        )
    return notes


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def write_csv(notes: Iterable[NoteRecord], path: Path) -> None:
    fieldnames = list(asdict(next(iter([NoteRecord('', '', '', '', '', '', '', '', 0, 0)]))).keys())
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for note in notes:
            writer.writerow(asdict(note))


def write_jsonl(notes: Iterable[NoteRecord], path: Path) -> None:
    with path.open("w", encoding="utf-8") as f:
        for note in notes:
            f.write(json.dumps(asdict(note), ensure_ascii=False) + "\n")


def write_markdown(notes: List[NoteRecord], path: Path) -> None:
    with path.open("w", encoding="utf-8") as f:
        f.write("# Apple Notes Export\n\n")
        f.write(f"Exported: {datetime.now().isoformat()}\n\n")
        f.write(f"Total notes: {len(notes)}\n\n")
        for idx, note in enumerate(notes, start=1):
            f.write(f"---\n\n## {idx}. {note.title or '(Untitled)'}\n\n")
            f.write(f"- Account: {note.account}\n")
            f.write(f"- Folder: {note.folder}\n")
            f.write(f"- ID: {note.id}\n")
            f.write(f"- Created: {note.created}\n")
            f.write(f"- Modified: {note.modified}\n")
            f.write(f"- Words: {note.word_count}\n\n")
            if note.body_text:
                f.write(note.body_text)
                f.write("\n\n")


def write_xlsx(notes: List[NoteRecord], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is not installed. Create a virtual environment and run: "
            "python -m pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "notes"

    headers = [
        "account",
        "folder",
        "id",
        "title",
        "created",
        "modified",
        "word_count",
        "char_count",
        "body_text",
        "body_html",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for n in notes:
        ws.append([
            n.account,
            n.folder,
            n.id,
            n.title,
            n.created,
            n.modified,
            n.word_count,
            n.char_count,
            n.body_text,
            n.body_html,
        ])

    widths = {
        "A": 18,
        "B": 24,
        "C": 26,
        "D": 36,
        "E": 24,
        "F": 24,
        "G": 12,
        "H": 12,
        "I": 100,
        "J": 80,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        row[8].alignment = Alignment(wrap_text=True, vertical="top")
        row[9].alignment = Alignment(wrap_text=True, vertical="top")

    summary = wb.create_sheet("summary")
    summary.append(["metric", "value"])
    summary.append(["exported_at", datetime.now().isoformat()])
    summary.append(["total_notes", len(notes)])
    summary.append(["accounts", len({n.account for n in notes})])
    summary.append(["folders", len({(n.account, n.folder) for n in notes})])
    summary.append(["total_words", sum(n.word_count for n in notes)])
    for cell in summary[1]:
        cell.font = Font(bold=True)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 32

    wb.save(path)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Export Apple Notes into merged files.")
    p.add_argument("--output-dir", required=True, help="Directory for export outputs")
    p.add_argument("--account", default=None, help="Optional exact account name filter, e.g. iCloud")
    p.add_argument("--skip-xlsx", action="store_true", help="Skip XLSX generation")
    return p.parse_args()


def main() -> int:
    if sys.platform != "darwin":
        print("This script only runs on macOS.", file=sys.stderr)
        return 2

    args = parse_args()
    out_dir = Path(os.path.expanduser(args.output_dir)).resolve()
    ensure_dir(out_dir)

    notes = fetch_notes(account_filter=args.account)
    if not notes:
        print("No notes found for the selected scope.")
        return 0

    csv_path = out_dir / "notes_export.csv"
    jsonl_path = out_dir / "notes_export.jsonl"
    md_path = out_dir / "notes_merged.md"
    xlsx_path = out_dir / "notes_export.xlsx"

    write_csv(notes, csv_path)
    write_jsonl(notes, jsonl_path)
    write_markdown(notes, md_path)
    if not args.skip_xlsx:
        write_xlsx(notes, xlsx_path)

    print(f"Export complete: {len(notes)} notes")
    print(f"CSV:   {csv_path}")
    print(f"JSONL: {jsonl_path}")
    print(f"MD:    {md_path}")
    if not args.skip_xlsx:
        print(f"XLSX:  {xlsx_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
